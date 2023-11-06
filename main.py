import openpyxl
import json
path = "city.xlsx"
wb_obj = openpyxl.load_workbook(path)
 
sheet_obj = wb_obj.active
 
max_col = sheet_obj.max_row
 
last_region = None
# Loop will print all columns name

json_data = {}

for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1) 
    if cell_obj.value is not None:
        last_region = cell_obj.value
    
    if json_data.get(last_region) is None:
        json_data[last_region] = []

    cell_obj_2 = sheet_obj.cell(row=i, column = 2)
    json_data[last_region].append(cell_obj_2.value)

with open("file.json", "w+", encoding='utf8') as file:
    file.write(json.dumps(json_data, ensure_ascii=False))
    



    
